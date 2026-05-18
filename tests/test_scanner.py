"""Tests for parser corpus scanning."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from shutil import copy2

import openpyxl
from openpyxl.styles import PatternFill

from bestrong.parser.scanner import (
    _cell_text,
    _classify_cell,
    _clip,
    _format_table,
    _harvest_exercise_vocabulary,
    _summarize_sheet,
    scan_corpus,
    write_scan_report,
)


FIXTURES = Path(__file__).parent / "fixtures"
CORPUS_FILES = [
    "athlete_a_program_38.xlsx",
    "athlete_b_program_35.xlsx",
]


def _copy_corpus(tmp_path: Path) -> Path:
    staging = tmp_path / "staging"
    staging.mkdir()
    for name in CORPUS_FILES:
        copy2(FIXTURES / name, staging / name)
    return staging


def test_scan_corpus_reports_fixture_sheet_inventory_and_vocabulary(tmp_path):
    staging = _copy_corpus(tmp_path)

    report = scan_corpus(staging)

    assert report.startswith("# Parser Scan")
    assert "- Workbooks scanned: **2**" in report
    assert "- Open errors: **0**" in report
    assert "| `HomePage` | 2 / 2 |" in report
    assert "| `Strength Block` | 2 / 2 |" in report
    assert "| `Warm Up Reps Calc` | 1 / 2 |" in report
    assert "### `Strength Block`" in report
    assert "**Column census** (first workbook's instance)" in report
    assert "- `Competition Bench Press 0-1-0` (44)" in report
    assert "- `Competition Squat` (20)" in report
    assert "- `Sumo Deadlift` (12)" in report


def test_write_scan_report_writes_default_report_path(tmp_path):
    staging = _copy_corpus(tmp_path)

    out_path = write_scan_report(staging)

    assert out_path == staging / "SCAN.md"
    assert out_path.exists()
    text = out_path.read_text()
    assert text.startswith("# Parser Scan")
    assert "- Workbooks scanned: **2**" in text


def test_write_scan_report_writes_explicit_report_path(tmp_path):
    staging = _copy_corpus(tmp_path)
    out_path = tmp_path / "reports" / "SCAN.md"

    result = write_scan_report(staging, out_path)

    assert result == out_path
    assert out_path.exists()
    assert "## Sheet/tab inventory" in out_path.read_text()


def test_cell_text_normalizes_empty_datetime_and_strings():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = None
    ws["A2"] = "  Competition Squat  "
    ws["A3"] = datetime(2026, 5, 17, 9, 30)

    assert _cell_text(ws["A1"]) == ""
    assert _cell_text(ws["A2"]) == "Competition Squat"
    assert _cell_text(ws["A3"]) == "2026-05-17"


def test_classify_cell_distinguishes_blank_numeric_and_string():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = None
    ws["A2"] = "   "
    ws["A3"] = 405
    ws["A4"] = "Bench"

    assert _classify_cell(ws["A1"]) == "blank"
    assert _classify_cell(ws["A2"]) == "blank"
    assert _classify_cell(ws["A3"]) == "numeric"
    assert _classify_cell(ws["A4"]) == "string"


def test_clip_replaces_markdown_breakers_and_truncates():
    assert _clip("a|b\nc", 20) == "a/b c"
    assert _clip("abcdef", 4) == "abc\u2026"


def test_format_table_pads_rows_and_clips_cells():
    table = _format_table(
        [
            ["Name", "Description"],
            ["A|B"],
            ["Lift", "abcdefghijklmnopqrstuvwxyz"],
        ]
    )

    assert "| Name | Description |" in table
    assert "| --- | --- |" in table
    assert "| A/B |  |" in table
    assert "abcdefghijklmnopqrstuvw\u2026" in table


def test_format_table_handles_empty_rows():
    assert _format_table([]) == "_(empty)_"


def test_summarize_sheet_collects_structure_and_samples():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grid"
    ws.sheet_state = "hidden"
    ws.append(["Movement", "Sets", "Notes"])
    ws.append(["Competition Squat", 3, "smooth"])
    ws.append(["Paused Bench", 4, "fast"])
    ws.append(["", None, ""])
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor="FFEA9999")
    ws.merge_cells("D1:E1")

    summary = _summarize_sheet("fixture.xlsx", ws)

    assert summary.name == "Grid"
    assert summary.workbook == "fixture.xlsx"
    assert summary.rows == 4
    assert summary.cols == 5
    assert summary.hidden is True
    assert summary.merged_ranges == 1
    assert summary.header_candidates[0][0] == 1
    assert summary.header_candidates[0][1][:3] == ["Movement", "Sets", "Notes"]
    assert summary.columns[0].string == 3
    assert summary.columns[1].numeric == 2
    assert summary.fill_colors["FFEA9999"] == 1
    assert summary.sample_rows[1][:3] == ["Competition Squat", "3", "smooth"]


def test_harvest_exercise_vocabulary_from_string_heavy_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    for value in [
        "Competition Squat",
        "Paused Bench",
        "Sumo Deadlift",
        "Target RPE",
        "Competition Squat",
    ]:
        ws.append([value])
    summary = _summarize_sheet("vocab.xlsx", ws)

    vocab = _harvest_exercise_vocabulary([summary])

    assert vocab["Competition Squat"] == 2
    assert vocab["Paused Bench"] == 1
    assert vocab["Sumo Deadlift"] == 1

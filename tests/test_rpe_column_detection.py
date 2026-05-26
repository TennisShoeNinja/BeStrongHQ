"""Per-tab RPE column detection across layout eras.

The actual-RPE column ("RPE Last Set") must be located by its own header
content, not assumed to sit one column right of "Target RPE". Older BeStrong
program templates reorder the band, e.g. inserting a backoff-weight column
between the prescription and the logged RPE. The previous adapter derived
``actual_rpe`` as ``target_rpe + 1`` and a fixed ``volume`` offset, so on those
layouts it read a weight (200-435 lbs) into the RPE field, which then got
flagged ``rpe_needs_review`` (the Jonathan Eppler December-2023 workbooks).

These tests lock in header-based detection and the standard layout's offsets so
the bundled-template behaviour never regresses.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import PatternFill

from bestrong.parser.adapters.bestrong_parser import BeStrongParser, _detect_week_bands

# Bench-green fill the adapter uses to classify a competition bench lift.
_BENCH_FILL = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")


def _workbook(header_labels: list[str], data_values: list) -> openpyxl.Workbook:
    """Build a minimal two-band Strength Block workbook.

    ``header_labels``/``data_values`` are the cells written from the
    exercise-name column rightward (index 0 is the name column). Two week bands
    are emitted so band detection does not fall back to hardcoded columns.
    """
    wb = openpyxl.Workbook()
    hp = wb.active
    hp.title = "HomePage"
    hp.cell(row=9, column=2, value="Jonathan Eppler")
    hp.cell(row=4, column=6, value=500)
    hp.cell(row=4, column=7, value=320)
    hp.cell(row=4, column=8, value=600)
    hp.cell(row=4, column=9, value=1420)

    sb = wb.create_sheet("Strength Block")
    for name_col in (1, 9):
        sb.cell(row=5, column=name_col, value="Day 1")
        sb.cell(row=6, column=name_col, value="Monday")
        header_row = 7
        for off, label in enumerate(header_labels):
            sb.cell(row=header_row, column=name_col + off, value=label)
        data_row = header_row + 1
        for off, value in enumerate(data_values):
            sb.cell(row=data_row, column=name_col + off, value=value)
        sb.cell(row=data_row, column=name_col).fill = _BENCH_FILL
    return wb


_STANDARD_HEADER = ["Primary Bench", "Sets", "Reps", "Weight", "Target RPE", "RPE Last Set", "Volume"]

# Older era: a backoff-weight column sits between Target RPE and RPE Last Set,
# so the logged RPE is two columns right of the prescription, not one.
_BACKOFF_HEADER = ["Primary Bench", "Sets", "Reps", "Weight", "Target RPE", "Backoff Wt", "RPE Last Set", "Volume"]


def _first_band(wb: openpyxl.Workbook) -> dict:
    ws = wb[BeStrongParser._find_block_sheet(wb)]
    return _detect_week_bands(ws)[0]


def _competition_lift(wb: openpyxl.Workbook):
    program = BeStrongParser().extract(wb)
    for session in program.sessions:
        for ex in session.exercises:
            if not ex.is_accessory:
                return ex
    raise AssertionError("no competition lift parsed")


def test_standard_layout_offsets_unchanged():
    """The bundled template's contiguous layout keeps the canonical offsets."""
    band = _first_band(_workbook(_STANDARD_HEADER, ["Competition Bench Press", 1, 1, 315, "@ 7 RPE", 8, 315]))
    assert band["weight"] == 3
    assert band["target_rpe"] == 4
    assert band["actual_rpe"] == 5
    assert band["volume"] == 6


def test_backoff_layout_locates_rpe_by_header():
    """An inserted backoff-weight column must not push the RPE read onto it:
    actual_rpe is the 'RPE Last Set' column (offset 6), not target_rpe + 1."""
    band = _first_band(_workbook(_BACKOFF_HEADER, ["Competition Bench Press", 1, 1, 315, "@ 7 RPE", 275, 8, 315]))
    assert band["weight"] == 3
    assert band["target_rpe"] == 4
    assert band["actual_rpe"] == 6  # the real "RPE Last Set" column, not 5
    assert band["volume"] == 7


def test_backoff_layout_reads_real_rpe_not_weight():
    """End to end: the logged RPE (8) is read, not the backoff weight (275),
    so the competition top set is not falsely flagged for review."""
    ex = _competition_lift(_workbook(_BACKOFF_HEADER, ["Competition Bench Press", 1, 1, 315, "@ 7 RPE", 275, 8, 315]))
    assert ex.weight_lbs == 315.0
    assert ex.actual_rpe == 8.0
    assert ex.rpe_needs_review is False
    assert ex.rpe_raw_value is None
    assert ex.volume == 315.0


def test_standard_layout_reads_rpe_end_to_end():
    """The standard layout still reads the logged RPE correctly."""
    ex = _competition_lift(_workbook(_STANDARD_HEADER, ["Competition Bench Press", 1, 1, 315, "@ 7 RPE", 8, 315]))
    assert ex.weight_lbs == 315.0
    assert ex.actual_rpe == 8.0
    assert ex.rpe_needs_review is False

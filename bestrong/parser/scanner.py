"""Structural scanner for staged spreadsheets.

Reads every xlsx in a staging directory and emits a markdown report describing
the corpus shape: tab inventory, header row candidates, column census per
recurring tab, fill-color vocabulary, merged-cell quirks, and an exercise-name
vocabulary harvested from string-heavy columns. The report is the brief Alex
hands back to the assistant when building a per-instance adapter — full corpus
knowledge up front instead of one-example guessing.

This module deliberately depends on nothing in the rest of the parser package
so it can run before any adapter exists for a instance.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell


_MAX_HEADER_CANDIDATES = 5
_MAX_SAMPLE_ROWS = 12
_MAX_SAMPLE_COLS = 10
_MAX_VOCAB = 100
_MAX_DISTINCT_VALUES = 5


@dataclass
class ColumnStats:
    index: int
    letter: str
    numeric: int = 0
    string: int = 0
    blank: int = 0
    distinct_values: Counter = field(default_factory=Counter)

    @property
    def total(self) -> int:
        return self.numeric + self.string + self.blank


@dataclass
class SheetSummary:
    name: str
    workbook: str
    rows: int
    cols: int
    hidden: bool
    merged_ranges: int
    header_candidates: list[tuple[int, list[str]]]
    columns: list[ColumnStats]
    fill_colors: Counter
    sample_rows: list[list[str]]


def _cell_text(cell: Cell) -> str:
    if cell.value is None:
        return ""
    if isinstance(cell.value, datetime):
        return cell.value.strftime("%Y-%m-%d")
    return str(cell.value).strip()


def _classify_cell(cell: Cell) -> str:
    if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
        return "blank"
    if isinstance(cell.value, (int, float)):
        return "numeric"
    return "string"


def _summarize_sheet(workbook_name: str, ws) -> SheetSummary:
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    columns = [
        ColumnStats(index=i, letter=openpyxl.utils.get_column_letter(i))
        for i in range(1, cols + 1)
    ]
    fill_colors: Counter = Counter()
    header_score: list[tuple[int, int, list[str]]] = []

    for row_idx in range(1, rows + 1):
        labels: list[str] = []
        string_count = 0
        for col_idx in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            kind = _classify_cell(cell)
            stat = columns[col_idx - 1]
            if kind == "numeric":
                stat.numeric += 1
            elif kind == "string":
                stat.string += 1
                text = _cell_text(cell)
                stat.distinct_values[text] += 1
                labels.append(text)
                string_count += 1
            else:
                stat.blank += 1
                labels.append("")

            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                rgb = fill.fgColor.rgb
                if isinstance(rgb, str) and rgb not in ("FFFFFFFF",):
                    fill_colors[rgb] += 1

        if string_count >= 2:
            header_score.append((row_idx, string_count, labels[:_MAX_SAMPLE_COLS]))

    header_score.sort(key=lambda t: t[1], reverse=True)
    header_candidates = [(row_idx, labels) for row_idx, _, labels in header_score[:_MAX_HEADER_CANDIDATES]]

    sample_rows: list[list[str]] = []
    sample_limit = min(rows, _MAX_SAMPLE_ROWS)
    sample_col_limit = min(cols, _MAX_SAMPLE_COLS)
    for row_idx in range(1, sample_limit + 1):
        sample_rows.append([
            _cell_text(ws.cell(row=row_idx, column=col_idx))
            for col_idx in range(1, sample_col_limit + 1)
        ])

    return SheetSummary(
        name=ws.title,
        workbook=workbook_name,
        rows=rows,
        cols=cols,
        hidden=ws.sheet_state != "visible",
        merged_ranges=len(ws.merged_cells.ranges),
        header_candidates=header_candidates,
        columns=columns,
        fill_colors=fill_colors,
        sample_rows=sample_rows,
    )


def _harvest_exercise_vocabulary(summaries: list[SheetSummary]) -> Counter:
    """Pull distinct strings from columns that look like exercise-name columns.

    Heuristic: any column where strings outnumber numerics by >=2x and at
    least 5 string entries are present. Multi-token strings beat single
    tokens — most program sheets have a "Bench Press" / "Paused Squat"
    column rather than a single-word column.
    """
    vocab: Counter = Counter()
    for summary in summaries:
        for col in summary.columns:
            if col.string < 5:
                continue
            if col.string < col.numeric * 2:
                continue
            multi_token = sum(1 for v in col.distinct_values if len(v.split()) >= 2)
            if multi_token < 2:
                continue
            for value, count in col.distinct_values.items():
                if not value:
                    continue
                vocab[value] += count
    return vocab


def _format_table(rows: list[list[str]]) -> str:
    if not rows:
        return "_(empty)_"
    width = max(len(r) for r in rows)
    normalized = [r + [""] * (width - len(r)) for r in rows]
    header = normalized[0]
    sep = ["---"] * width
    body = normalized[1:]
    out = ["| " + " | ".join(_clip(c, 24) for c in header) + " |",
           "| " + " | ".join(sep) + " |"]
    for row in body:
        out.append("| " + " | ".join(_clip(c, 24) for c in row) + " |")
    return "\n".join(out)


def _clip(text: str, max_len: int) -> str:
    text = (text or "").replace("|", "/").replace("\n", " ")
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


def scan_corpus(staging_dir: Path) -> str:
    """Walk ``staging_dir`` for xlsx files and emit a markdown scan report."""
    files = sorted(staging_dir.glob("*.xlsx"))
    if not files:
        return f"# Parser Scan\n\nNo xlsx files found in `{staging_dir}`.\n"

    summaries: list[SheetSummary] = []
    sheet_name_counter: Counter = Counter()
    file_to_sheets: dict[str, list[str]] = defaultdict(list)
    open_errors: list[tuple[str, str]] = []

    for path in files:
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        except Exception as exc:
            open_errors.append((path.name, str(exc)))
            continue
        for ws in wb.worksheets:
            sheet_name_counter[ws.title] += 1
            file_to_sheets[path.name].append(ws.title)
            summaries.append(_summarize_sheet(path.name, ws))
        wb.close()

    grouped: dict[str, list[SheetSummary]] = defaultdict(list)
    for s in summaries:
        grouped[s.name].append(s)

    out: list[str] = []
    out.append("# Parser Scan")
    out.append("")
    out.append(f"_Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} from `{staging_dir}`_")
    out.append("")
    out.append(f"- Workbooks scanned: **{len(files)}**")
    out.append(f"- Distinct sheet/tab names: **{len(sheet_name_counter)}**")
    out.append(f"- Open errors: **{len(open_errors)}**")
    out.append("")

    if open_errors:
        out.append("## Open errors")
        for name, err in open_errors:
            out.append(f"- `{name}`: {_clip(err, 120)}")
        out.append("")

    out.append("## Sheet/tab inventory")
    out.append("")
    out.append("| Tab name | Workbooks present in |")
    out.append("| --- | --- |")
    for name, count in sheet_name_counter.most_common():
        out.append(f"| `{name}` | {count} / {len(files)} |")
    out.append("")


    set_counter: Counter = Counter()
    for fname, sheets in file_to_sheets.items():


        set_counter[tuple(sorted(sheets))] += 1


    def _normalize_set(names: tuple[str, ...]) -> tuple[str, ...]:
        return tuple(sorted("".join(n.lower().split()) for n in names))

    norm_groups: dict[tuple[str, ...], list[tuple[tuple[str, ...], int]]] = defaultdict(list)
    for sheet_set, count in set_counter.items():
        norm_groups[_normalize_set(sheet_set)].append((sheet_set, count))

    out.append("## Distinct sheet-name sets")
    out.append("")
    out.append(
        f"Each row is a unique combination of sheet names that appear together "
        f"in a workbook. **{len(set_counter)}** distinct set(s) across "
        f"{len(files)} workbook(s)."
    )
    out.append("")
    near_dupe_groups = [g for g in norm_groups.values() if len(g) > 1]
    if near_dupe_groups:
        out.append(
            f"WARN: **{len(near_dupe_groups)} near-duplicate group(s)** detected "
            f"(case- and whitespace-insensitive). Likely a sheet-name variant "
            f"that ``can_parse`` needs to accept; check before writing the adapter."
        )
        out.append("")
    out.append("| Workbooks | Sheet set |")
    out.append("| --- | --- |")
    for sheet_set, count in sorted(set_counter.items(), key=lambda kv: -kv[1]):
        normalized = _normalize_set(sheet_set)
        marker = " (!)" if len(norm_groups[normalized]) > 1 else ""
        rendered = ", ".join(f"`{n}`" for n in sheet_set)
        out.append(f"| {count}{marker} | {rendered} |")
    out.append("")
    if near_dupe_groups:
        out.append("**Near-duplicate groups** (rows above marked (!)):")
        out.append("")
        for group in near_dupe_groups:
            variants = " vs. ".join(
                "{" + ", ".join(f"`{n}`" for n in s) + "}"
                for s, _ in sorted(group, key=lambda x: -x[1])
            )
            out.append(f"- {variants}")
        out.append("")

    out.append("## Per-tab detail")
    out.append("")
    for tab_name, group in sorted(grouped.items(), key=lambda kv: -len(kv[1])):
        out.append(f"### `{tab_name}`")
        out.append("")
        rows_avg = sum(s.rows for s in group) / len(group)
        cols_avg = sum(s.cols for s in group) / len(group)
        hidden_count = sum(1 for s in group if s.hidden)
        merged_count = sum(s.merged_ranges for s in group)
        out.append(
            f"- Appears in {len(group)} workbook(s); "
            f"avg dims {rows_avg:.0f} rows × {cols_avg:.0f} cols; "
            f"hidden in {hidden_count}; "
            f"merged-cell ranges total {merged_count}"
        )


        header_freq: Counter = Counter()
        for s in group:
            for row_idx, labels in s.header_candidates:
                header_freq[(row_idx, tuple(labels))] += 1
        if header_freq:
            out.append("")
            out.append("**Header row candidates** (row index → labels, frequency)")
            out.append("")
            for (row_idx, labels), freq in header_freq.most_common(_MAX_HEADER_CANDIDATES):
                preview = " · ".join(_clip(c, 18) for c in labels if c) or "_(blanks)_"
                out.append(f"- row {row_idx} ({freq}×): {preview}")


        first = group[0]
        out.append("")
        out.append("**Column census** (first workbook's instance)")
        out.append("")
        rows = [["Col", "% num", "% str", "% blank", "Top values"]]
        for col in first.columns:
            total = max(col.total, 1)
            top = ", ".join(
                f"{_clip(v, 14)}({c})"
                for v, c in col.distinct_values.most_common(_MAX_DISTINCT_VALUES)
                if v
            )
            rows.append([
                col.letter,
                f"{col.numeric * 100 / total:.0f}",
                f"{col.string * 100 / total:.0f}",
                f"{col.blank * 100 / total:.0f}",
                top or "—",
            ])
        out.append(_format_table(rows))


        out.append("")
        out.append(f"**Sample rows** (`{first.workbook}`, first {len(first.sample_rows)} rows)")
        out.append("")
        out.append(_format_table(first.sample_rows))


        agg_colors: Counter = Counter()
        for s in group:
            agg_colors.update(s.fill_colors)
        if agg_colors:
            out.append("")
            out.append("**Fill colors** (rgb → cell count across all instances)")
            out.append("")
            for rgb, count in agg_colors.most_common(15):
                out.append(f"- `#{rgb}` — {count} cells")
        out.append("")

    vocab = _harvest_exercise_vocabulary(summaries)
    if vocab:
        out.append("## Likely exercise-name vocabulary")
        out.append("")
        out.append(
            "Strings harvested from columns where strings dominate numerics ≥2× "
            "and ≥2 multi-token entries appear. Frequency-sorted across the "
            "whole corpus."
        )
        out.append("")
        for value, count in vocab.most_common(_MAX_VOCAB):
            out.append(f"- `{_clip(value, 80)}` ({count})")
        out.append("")

    out.append("## What to fill in next")
    out.append("")
    out.append(
        "Hand this report to the assistant alongside the parser-brief template "
        "at `.mex/patterns/parser_brief.md`. The brief asks for the things the "
        "scan can't decide for you — which tab is the program grid, which row "
        "is the canonical header, what each color means, and how the coach "
        "expects rep / weight / RPE to be split across columns."
    )
    out.append("")

    return "\n".join(out)


def write_scan_report(staging_dir: Path, out_path: Path | None = None) -> Path:
    """Write the scan report next to the staging dir (or to ``out_path``).

    Returns the resolved output path. Default: ``<staging_dir>/SCAN.md``.
    """
    report = scan_corpus(staging_dir)
    target = out_path or (staging_dir / "SCAN.md")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(report)
    return target

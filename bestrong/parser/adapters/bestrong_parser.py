"""Default BeStrong parser for the included spreadsheet templates.

Handles the "Strength Block" layout shipped in templates/bestrong_template_4day.xlsx
and bestrong_template_5day.xlsx, where:
- Weeks are arranged horizontally in column bands
- Days are stacked vertically within each band
- Each day has a pre-session assessment block followed by exercise data
"""

from __future__ import annotations

import re
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ...utils.exercise_names import name_indicates_accessory
from ..rpe_cell import parse_rpe_cell
from . import (
    AthleteInfo,
    BaseAdapter,
    ExerciseRow,
    ProgramData,
    SessionData,
    SessionInfo,
    WellnessEntry,
    register,
)


_LIFT_PATTERNS: dict[str, list[str]] = {
    "squat": [
        "competition squat",
        "comp squat",
        "squat",
    ],
    "bench": [
        "competition bench",
        "comp bench",
        "close grip bench",
        "bench press",
        "bench",
    ],
    "deadlift": [
        "sumo deadlift",
        "sumo cluster",
        "conventional deadlift",
        "deadlift",
        "rdl",
        "paused sumo",
        "deficit deadlift",
        "block pull",
    ],
}


def classify_lift(exercise_name: str) -> str:
    """Classify an exercise into squat/bench/deadlift/accessory by name pattern."""
    name_lower = exercise_name.lower()
    for category, patterns in _LIFT_PATTERNS.items():
        for pattern in patterns:
            if pattern in name_lower:
                return category
    return "accessory"


# ``name_indicates_accessory`` and its token list moved to
# ``bestrong.utils.exercise_names`` so PR logging can share the same
# accessory vocabulary; imported above and re-exported here for callers
# (and tests) that reference it on this module.


_COLOR_MAP: dict[str, tuple[int, int, int]] = {
    "bench": (0xD9, 0xEA, 0xD3),
    "squat": (0xFF, 0xF2, 0xCC),
    "deadlift": (0xC9, 0xDA, 0xF8),
}


_BRIGHT_YELLOW = (0xFF, 0xFF, 0x00)


_ATHLETE_FILL_WEIGHT = (0xB6, 0xD7, 0xA8)
_ATHLETE_FILL_RPE = (0xEA, 0x99, 0x99)


_COLOR_TOLERANCE = 20


def _rgb_distance(c1: tuple[int, int, int], c2: tuple[int, int, int]) -> int:
    """Manhattan distance between two RGB colors."""
    return abs(c1[0] - c2[0]) + abs(c1[1] - c2[1]) + abs(c1[2] - c2[2])


def _get_cell_bg_color(ws: Worksheet, row: int, col: int) -> tuple[int, int, int] | None:
    """Extract the background RGB color of a cell, or None if no fill."""
    cell = ws.cell(row=row, column=col)
    fill = cell.fill


    if fill.fill_type is None or fill.fill_type == "none":
        return None

    fg = fill.fgColor
    if fg is None:
        return None


    if fg.type == "rgb" and fg.rgb:
        hex_str = str(fg.rgb)


        if hex_str in ("00000000", "FFFFFFFF"):
            return None
        if len(hex_str) == 8:
            r = int(hex_str[2:4], 16)
            g = int(hex_str[4:6], 16)
            b = int(hex_str[6:8], 16)
            return (r, g, b)


    return None


def is_weight_expected_cell(ws: Worksheet, row: int, col: int) -> bool:
    """True when the weight cell is painted Alex's athlete-fill green,
    meaning the coach is asking the athlete to fill in a weight.

    Paired with ``rpe_expected`` in the data-quality pipeline — a row can
    have either, both, or neither. Gray (coach-prescribed) weights return
    False; that's a separate signal that the weight was dictated.
    """
    rgb = _get_cell_bg_color(ws, row, col)
    if rgb is None:
        return False
    return _rgb_distance(rgb, _ATHLETE_FILL_WEIGHT) < _COLOR_TOLERANCE


def is_rpe_expected_cell(ws: Worksheet, row: int, col: int) -> bool:
    """True when the RPE Last Set cell is painted Alex's athlete-fill pink,
    meaning the coach expects an RPE report from the athlete on this row.

    Drives the RPE-compliance denominator: only rows where this is True
    count toward compliance. Backdowns without a pink cell are invisible
    to the compliance metric — the coach explicitly didn't ask.
    """
    rgb = _get_cell_bg_color(ws, row, col)
    if rgb is None:
        return False
    return _rgb_distance(rgb, _ATHLETE_FILL_RPE) < _COLOR_TOLERANCE


def classify_by_color(ws: Worksheet, row: int, col: int) -> str | None:
    """Classify an exercise by its cell background color.

    Returns "squat", "bench", "deadlift", or None (meaning no color /
    accessory / could not determine).
    """
    rgb = _get_cell_bg_color(ws, row, col)
    if rgb is None:
        return None


    if _rgb_distance(rgb, _BRIGHT_YELLOW) < _COLOR_TOLERANCE:
        return None


    best_match: str | None = None
    best_dist = _COLOR_TOLERANCE + 1

    for lift, ref_color in _COLOR_MAP.items():
        dist = _rgb_distance(rgb, ref_color)
        if dist < best_dist:
            best_dist = dist
            best_match = lift

    return best_match


_FALLBACK_NAME_COLS = [1, 9, 18, 26]


_STANDARD_OFFSETS = {"sets": 1, "reps": 2, "weight": 3, "target_rpe": 4, "actual_rpe": 5, "volume": 6}


# How far right of the exercise-name column to look for the header band. The
# widest layout seen carries a blank spacer column plus the six labelled
# columns, so a 12-wide window covers every era without straying into the
# next week's band.
_BAND_HEADER_WIDTH = 12


# RPE-header vocabulary. The team's actual-RPE column is headed "RPE Last Set"
# (older eras also "Actual RPE" / "Last Set RPE" / "RPE Achieved"); the target
# column always carries a prescription word. A bare "RPE" with no qualifier is
# left for positional resolution (the target column always sits left of the
# actual-RPE column).
_TARGET_RPE_WORDS = ("target", "goal", "prescrib")
_ACTUAL_RPE_WORDS = ("last", "actual", "achiev", "got", "done")


def _band_offsets_from_header(ws: Worksheet, name_col: int, header_row: int) -> dict | None:
    """Map a band's columns to fields from one header row, by header content.

    Returns an offset dict (``sets``/``reps``/``weight``/``target_rpe``/
    ``actual_rpe``/``volume``, each relative to ``name_col``), or ``None`` when
    ``header_row`` is not an exercise header (no "Weight" label in the band).

    The actual-RPE column is located by its *own* header text rather than
    assumed to sit one column right of "Target RPE". Older layout eras shift or
    reorder the RPE columns, and the previous offset-from-target guess silently
    pointed the RPE read at a weight or volume column, which is the source of
    the spurious ``rpe_needs_review`` weights (200-435 lbs) on athletes with
    vintage workbooks. Any field whose header is absent falls back to
    ``_STANDARD_OFFSETS``, with the historical adjacency guesses kept only as a
    last resort.
    """
    sets_off = reps_off = weight_off = volume_off = None
    target_off = actual_off = None
    bare_rpe_offs: list[int] = []

    for off in range(1, _BAND_HEADER_WIDTH + 1):
        val = ws.cell(row=header_row, column=name_col + off).value
        if not isinstance(val, str):
            continue
        low = val.strip().lower()
        if not low:
            continue
        if low == "sets" and sets_off is None:
            sets_off = off
        elif low == "reps" and reps_off is None:
            reps_off = off
        elif low == "weight" and weight_off is None:
            weight_off = off
        elif low.startswith("volume") and volume_off is None:
            volume_off = off
        elif "rpe" in low:
            if any(w in low for w in _TARGET_RPE_WORDS):
                if target_off is None:
                    target_off = off
            elif any(w in low for w in _ACTUAL_RPE_WORDS):
                if actual_off is None:
                    actual_off = off
            else:
                bare_rpe_offs.append(off)

    # "Weight" is the one header every exercise era carries; without it this
    # row is the pre-session block, a day-split table, or just noise.
    if weight_off is None:
        return None

    # Resolve any unqualified "RPE" headers positionally: target sits left of
    # the actual-RPE column.
    leftover = [o for o in bare_rpe_offs if o not in (target_off, actual_off)]
    if target_off is None and leftover:
        target_off = min(leftover)
        leftover = [o for o in leftover if o != target_off]
    if actual_off is None and leftover:
        actual_off = max(leftover)

    # Adjacency fallbacks for layouts exposing only one of the two RPE headers
    # (the historical assumption), kept strictly as a last resort.
    if target_off is not None and actual_off is None:
        actual_off = target_off + 1
    if actual_off is not None and target_off is None:
        target_off = actual_off - 1
    if actual_off is not None and volume_off is None:
        volume_off = actual_off + 1

    offsets = dict(_STANDARD_OFFSETS)
    for field, off in (
        ("sets", sets_off),
        ("reps", reps_off),
        ("weight", weight_off),
        ("target_rpe", target_off),
        ("actual_rpe", actual_off),
        ("volume", volume_off),
    ):
        if off is not None:
            offsets[field] = off
    return offsets


def _detect_week_bands(ws: Worksheet) -> list[dict]:
    """Detect the column band for each week by scanning for Day markers.

    Scans the sheet for "Day N" markers (any day number) to find which columns
    contain exercise names for each week. Some weeks may lack a "Day 1" marker
    (e.g., one fixture's Week 4) so we look for *any* "Day \\d" pattern, then
    read each band's column offsets from its header row by content (see
    ``_band_offsets_from_header``). Falls back to hardcoded name columns and
    standard offsets if detection fails.
    """
    _DAY_RE = re.compile(r"^Day\s+\d+$")


    day_cols: set[int] = set()
    for c in range(1, min(ws.max_column + 1, 40)):
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and _DAY_RE.match(v.strip()):
                day_cols.add(c)
                break


    sorted_cols = sorted(day_cols)

    if len(sorted_cols) < 2:

        sorted_cols = _FALLBACK_NAME_COLS


    bands: list[dict] = []
    for name_col in sorted_cols:
        offsets = dict(_STANDARD_OFFSETS)

        # The header band repeats per session down a week column but at fixed
        # offsets, so the first "Weight"-bearing header row fully describes the
        # band's layout era.
        for r in range(1, min(ws.max_row + 1, 100)):
            detected = _band_offsets_from_header(ws, name_col, r)
            if detected is not None:
                offsets = detected
                break

        bands.append({"name_col": name_col, **offsets})

    return bands


_TITLE_RE = re.compile(
    r"^(.+?)(?:'s|'s|\u2019s)\s+Program\s+(\d+)"
    r"\s*[–—\-]\s*"
    r"\(\s*(\d{1,2}[/_ ]\d{1,2}[/_ ]\d{2,4})"
    r"\s*[–—\-]\s*"
    r"(\d{1,2}[/_ ]\d{1,2}[/_ ]\d{2,4})\s*\)"
    r"\s*[–—\-]\s*"
    r"(.+)$"
)


_OLD_TITLE_RE = re.compile(
    r"^(.+?)(?:'s|'s|\u2019s)\s+Program\s*"
    r"\(?\s*(\d{1,2}[/_ ]\d{1,2}[/_ ]\d{2,4})"
    r"\s*[–—\-]\s*"
    r"(\d{1,2}[/_ ]\d{1,2}[/_ ]\d{2,4})\s*\)?"
    r"\s*[–—\-]?\s*"
    r"(.*)$"
)


_NUMBER_RE = re.compile(r"Program\s+(\d+)", re.IGNORECASE)


_ATHLETE_RE = re.compile(r"^(.+?)(?:'s|'s|\u2019s)\s+", re.IGNORECASE)


def parse_filename(filename: str) -> dict[str, Any]:
    """Extract program metadata from a filename or title string.

    Handles the standard title format:
        Ed's Program 35 – (02/01/26 – 02/22/26) – Strength Block

    Also handles mangled filenames from Drive downloads by cleaning underscores first.
    """

    clean = re.sub(r"\.xlsx$", "", filename, flags=re.IGNORECASE).strip()

    clean = clean.replace("____", " – ").replace("__", " ").replace("_", " ").strip()

    result: dict[str, Any] = {
        "athlete_name": None,
        "program_number": None,
        "program_name": None,
        "date_start": None,
        "date_end": None,
    }


    m = _TITLE_RE.match(clean)
    if m:
        result["athlete_name"] = m.group(1).strip()
        result["program_number"] = int(m.group(2))

        result["date_start"] = re.sub(r"[_ ]", "/", m.group(3))
        result["date_end"] = re.sub(r"[_ ]", "/", m.group(4))
        result["program_name"] = m.group(5).strip()
        return result


    m = _OLD_TITLE_RE.match(clean)
    if m:
        result["athlete_name"] = m.group(1).strip()
        result["date_start"] = re.sub(r"[_ ]", "/", m.group(2))
        result["date_end"] = re.sub(r"[_ ]", "/", m.group(3))
        rest = m.group(4).strip()
        result["program_name"] = rest if rest else clean

        num_m = _NUMBER_RE.search(clean)
        if num_m:
            result["program_number"] = int(num_m.group(1))
        return result


    m = _NUMBER_RE.search(clean)
    if m:
        result["program_number"] = int(m.group(1))


    m = _ATHLETE_RE.match(clean)
    if m:
        result["athlete_name"] = m.group(1).strip()


    result["program_name"] = clean

    return result


def _cell_val(ws: Worksheet, row: int, col: int) -> Any:
    """Get cell value, returning None for empty cells."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _is_rpe_string(value: Any) -> bool:
    """Check if a value is an RPE prescription string (e.g., '@ 7 RPE')."""
    if not isinstance(value, str):
        return False
    return "@" in value and "RPE" in value.upper()


def _parse_rpe_actual(value: Any) -> tuple[float | None, bool]:
    """Parse an actual RPE value from a cell.

    Returns (rpe_value, failed) where failed is True when the cell
    contains 'fail' or 'failed' (case-insensitive).

    RPE values must be between 1 and 10 (inclusive). Values outside this
    range are almost certainly weight or volume numbers that landed in the
    wrong column due to a layout offset mismatch, so we discard them.
    """
    if value is None:
        return None, False
    if isinstance(value, (int, float)):
        v = float(value)
        if v < 1 or v > 10:
            return None, False
        return v, False
    if isinstance(value, str):
        s = value.strip()
        if s.lower() in ("fail", "failed"):
            return None, True
        if s in ("RPE?", "", "-"):
            return None, False
        try:
            v = float(s)
            if v < 1 or v > 10:
                return None, False
            return v, False
        except ValueError:
            return None, False
    return None, False


def _parse_weight(value: Any) -> tuple[float | None, bool, str | None]:
    """Parse a weight cell value.

    Returns: (weight_lbs, is_accessory, target_rpe_from_weight)
    """
    if value is None:
        return None, False, None

    if isinstance(value, (int, float)):
        return float(value), False, None

    if isinstance(value, str):
        s = value.strip()


        if _is_rpe_string(s):
            return None, True, s


        if s in ("??", "#VALUE!", "#REF!"):
            return None, False, None


        if "b/w" in s.lower() or "bodyweight" in s.lower():
            return None, True, s


        try:
            return float(s), False, None
        except ValueError:
            return None, False, None

    return None, False, None


def _parse_reps(value: Any) -> int | None:
    """Parse a reps value, handling edge cases."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):

        try:
            return int(value)
        except ValueError:
            return None
    return None


def _find_day_markers(ws: Worksheet, name_col: int) -> list[dict]:
    """Scan for day markers and return their positions.

    First looks for "Day N" markers in the name_col. If Day 1 is missing,
    falls back to "Week N DAYNAME" markers in the adjacent sets column
    (name_col + 1) to discover day boundaries.

    Returns list of dicts with keys: day_number, label_row
    """

    markers = []
    for row in range(1, ws.max_row + 1):
        val = _cell_val(ws, row, name_col)
        if val and isinstance(val, str) and val.strip().startswith("Day "):
            try:
                day_num = int(val.strip().split()[1])
            except (IndexError, ValueError):
                continue
            markers.append({"day_number": day_num, "label_row": row})


    day_nums_found = {m["day_number"] for m in markers}
    if 1 not in day_nums_found:
        _WEEK_DAY_RE = re.compile(r"^Week\s+\d+\s+\w+", re.IGNORECASE)
        week_markers = []
        for row in range(1, ws.max_row + 1):
            val = _cell_val(ws, row, name_col + 1)
            if val and isinstance(val, str) and _WEEK_DAY_RE.match(val.strip()):
                week_markers.append(row)

        if week_markers:


            markers = []
            for i, row in enumerate(sorted(week_markers)):
                markers.append({"day_number": i + 1, "label_row": row})

    return markers


def _find_day_name_and_header(
    ws: Worksheet, label_row: int, name_col: int, band: dict
) -> tuple[str, int]:
    """Find the day name and exercise header row starting from the day label row.

    Scans downward from label_row looking for the day name (Monday/Tuesday/etc)
    and the header row (contains 'Sets' in the sets column).
    """
    day_name = ""
    header_row = label_row + 5

    _DAY_NAMES = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}
    for r in range(label_row, label_row + 10):
        val = _cell_val(ws, r, name_col)
        if val and isinstance(val, str) and val.strip().lower().rstrip() in _DAY_NAMES:
            day_name = val.strip().title()

        sets_val = _cell_val(ws, r, name_col + band["sets"])
        if sets_val and isinstance(sets_val, str) and sets_val.strip() == "Sets":
            header_row = r
            break

    return day_name, header_row


def _find_session_label(ws: Worksheet, header_row: int, name_col: int) -> str | None:
    """Get the session label from the header row (e.g., 'Secondary Deads')."""
    val = _cell_val(ws, header_row, name_col)
    if val and isinstance(val, str):
        s = val.strip()

        if s not in ("Sets", "Reps", "Weight") and not s.startswith("Day "):
            return s
    return None


def _extract_pre_session(
    ws: Worksheet, label_row: int, sets_col: int
) -> dict[str, str | None]:
    """Extract pre-session assessment ratings (nutrition/stress/sleep/fatigue)."""
    ratings: dict[str, str | None] = {
        "nutrition": None,
        "stress": None,
        "sleep": None,
        "fatigue": None,
    }


    field_map = {
        "Nutrition": "nutrition",
        "Stress": "stress",
        "Sleep": "sleep",
        "Fatigue": "fatigue",
    }
    for r in range(label_row + 1, label_row + 8):
        label = _cell_val(ws, r, sets_col)
        if label and isinstance(label, str):
            key = field_map.get(label.strip())
            if key:
                val = _cell_val(ws, r, sets_col + 1)
                if val is not None:
                    ratings[key] = str(val).strip()
    return ratings


def _extract_exercises(
    ws: Worksheet,
    header_row: int,
    name_col: int,
    band: dict,
    max_row: int,
) -> list[ExerciseRow]:
    """Extract exercise rows from a session block.

    Starts at header_row + 1 and continues until hitting another 'Day' marker,
    a volume summary row, or max_row.
    """
    exercises: list[ExerciseRow] = []
    exercise_order = 0
    exercise_group = 0
    prev_name: str | None = None
    in_empty_gap = False


    group_top_set_weight: float | None = None

    cols = {
        "name": name_col,
        "sets": name_col + band["sets"],
        "reps": name_col + band["reps"],
        "weight": name_col + band["weight"],
        "target_rpe": name_col + band["target_rpe"],
        "actual_rpe": name_col + band["actual_rpe"],
        "volume": name_col + band["volume"],
    }

    for row in range(header_row + 1, max_row + 1):
        name_val = _cell_val(ws, row, cols["name"])
        sets_val = _cell_val(ws, row, cols["sets"])


        if name_val and isinstance(name_val, str):
            s = name_val.strip()
            if s.startswith("Day "):
                break


        if name_val and isinstance(name_val, str):
            if name_val.strip().lower().rstrip() in (
                "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"
            ):
                break


        if name_val is None and sets_val is None:
            in_empty_gap = True
            continue


        if name_val is None or sets_val is None:
            continue


        if isinstance(sets_val, str) and sets_val.strip() in ("Sets", "Fatigue"):
            continue
        if isinstance(name_val, str) and name_val.strip().startswith("Week "):
            continue


        raw_name = str(name_val).strip()
        manual_top_set_override = False
        ts_match = re.match(r"^\s*ts\s*:\s*", raw_name, flags=re.IGNORECASE)
        if ts_match:
            manual_top_set_override = True
            ex_name = raw_name[ts_match.end():].strip()
        else:
            ex_name = raw_name


        try:
            sets = int(float(sets_val))
        except (ValueError, TypeError):
            continue


        reps = _parse_reps(_cell_val(ws, row, cols["reps"]))


        weight_raw = _cell_val(ws, row, cols["weight"])
        weight_lbs, is_accessory, rpe_from_weight = _parse_weight(weight_raw)


        weight_expected = is_weight_expected_cell(ws, row, cols["weight"])
        rpe_expected = is_rpe_expected_cell(ws, row, cols["actual_rpe"])


        failed = False
        rpe_needs_review = False
        rpe_raw_value: str | None = None
        if is_accessory:
            target_rpe = rpe_from_weight
            actual_rpe = None
        else:
            target_rpe_raw = _cell_val(ws, row, cols["target_rpe"])
            target_rpe = str(target_rpe_raw).strip() if target_rpe_raw else None
            rpe_cell_val = _cell_val(ws, row, cols["actual_rpe"])
            rpe_result = parse_rpe_cell(rpe_cell_val)
            actual_rpe = rpe_result.parsed
            failed = rpe_result.failed
            rpe_needs_review = rpe_result.needs_review
            rpe_raw_value = rpe_result.raw


        vol_raw = _cell_val(ws, row, cols["volume"])
        volume = None
        if isinstance(vol_raw, (int, float)):
            volume = float(vol_raw)


        color_category = classify_by_color(ws, row, cols["name"])


        if color_category is not None and name_indicates_accessory(ex_name):
            color_category = None

        if is_accessory:
            lift_category = "accessory"
        elif color_category is not None:
            lift_category = color_category
            is_accessory = False
        else:
            lift_category = "accessory"
            is_accessory = True


        new_group = ex_name != prev_name or in_empty_gap
        if new_group or is_accessory:
            exercise_group += 1
            group_top_set_weight = None
        in_empty_gap = False
        prev_name = ex_name
        exercise_order += 1


        if is_accessory:
            set_type = "accessory"
        elif manual_top_set_override:
            set_type = "top_set"
            if weight_lbs is not None:
                group_top_set_weight = weight_lbs
        elif target_rpe and (weight_expected or weight_lbs is not None):
            if group_top_set_weight is None:
                set_type = "top_set"
                if weight_lbs is not None:
                    group_top_set_weight = weight_lbs
            elif weight_lbs is not None and weight_lbs >= group_top_set_weight:
                set_type = "top_set"
                group_top_set_weight = weight_lbs
            else:
                set_type = "backdown"
        else:
            set_type = "backdown"

        exercises.append(ExerciseRow(
            exercise_name=ex_name,
            exercise_order=exercise_order,
            exercise_group=exercise_group,
            set_type=set_type,
            lift_category=lift_category,
            sets=sets,
            reps=reps,
            weight_lbs=weight_lbs,
            target_rpe=target_rpe,
            actual_rpe=actual_rpe,
            volume=volume,
            is_accessory=is_accessory,
            failed=failed,
            notes=None,
            weight_expected=weight_expected,
            rpe_expected=rpe_expected,
            rpe_needs_review=rpe_needs_review,
            rpe_raw_value=rpe_raw_value,
        ))

    return exercises


_DAY_KEYWORDS = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}

_PRIMARY_LIFT_PATTERNS: dict[str, list[str]] = {
    "squat":    ["primary squat"],
    "bench":    ["primary bench"],
    "deadlift": ["primary deadlift", "primary dead"],
}


def _detect_primary_days(ws: "openpyxl.worksheet.worksheet.Worksheet") -> dict[str, int | None]:
    """Scan the top of the sheet for the weekly split table and return primary day numbers.

    Looks for a row containing 2+ day-name cells (Monday, Tuesday, etc.) within the
    first 30 rows. Maps each day-name column left-to-right to Day 1, Day 2, … then
    scans the next 8 rows for cells whose text starts with "Primary Squat/Bench/Deadlift"
    and records which day number that column corresponds to.

    Returns a dict with keys 'squat', 'bench', 'deadlift'. Values are 1-based day
    numbers or None when not found.
    """
    result: dict[str, int | None] = {"squat": None, "bench": None, "deadlift": None}

    header_row_idx: int | None = None
    col_to_day_num: dict[int, int] = {}


    for row_idx in range(1, 31):
        day_cols: list[int] = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and val.strip().lower() in _DAY_KEYWORDS:
                day_cols.append(col_idx)
        if len(day_cols) >= 2:
            header_row_idx = row_idx
            for i, col in enumerate(sorted(day_cols)):
                col_to_day_num[col] = i + 1
            break

    if not col_to_day_num:
        return result


    for row_idx in range(header_row_idx + 1, header_row_idx + 9):  # type: ignore[operator]
        if row_idx > ws.max_row:
            break
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if not isinstance(val, str):
                continue
            val_lower = val.strip().lower()
            for lift, patterns in _PRIMARY_LIFT_PATTERNS.items():
                if result[lift] is None:
                    for pat in patterns:
                        if val_lower.startswith(pat):
                            day_num = col_to_day_num.get(col_idx)
                            if day_num is not None:
                                result[lift] = day_num
                            break

    return result


@register
class BeStrongParser(BaseAdapter):
    """Default BeStrong parser for the included spreadsheet templates.

    Detects workbooks containing 'Strength Block' and 'HomePage' sheets
    with the standard horizontal-weeks, vertical-days layout.
    """

    name = "bestrong"
    description = "Default BeStrong spreadsheet (Strength Block layout)"


    _BLOCK_PREFIXES = [
        "Strength Block",
        "STRENGTH BLOCK",
        "Peaking Block",
        "PEAKING BLOCK",
        "Hypertrophy Block",
        "HYPERTROPHY BLOCK",
        "Block",
    ]

    @staticmethod
    def _find_block_sheet(workbook: openpyxl.Workbook) -> str | None:
        """Find the main training block sheet.

        Accepts variants like 'Strength Block', 'Strength Block 2',
        'PEAKING BLOCK', 'Hypertrophy Block', etc.
        """
        for prefix in BeStrongParser._BLOCK_PREFIXES:
            for name in workbook.sheetnames:
                if name.startswith(prefix):
                    return name
        return None

    @staticmethod
    def _find_strength_sheet(workbook: openpyxl.Workbook) -> str | None:
        """Find the main block sheet (legacy name kept for compatibility)."""
        return BeStrongParser._find_block_sheet(workbook)

    @staticmethod
    def can_parse(workbook: openpyxl.Workbook) -> bool:
        sheets = workbook.sheetnames
        has_homepage = "HomePage" in sheets
        has_block = BeStrongParser._find_block_sheet(workbook) is not None
        return has_block and has_homepage

    def extract(self, workbook: openpyxl.Workbook, filename: str = "") -> ProgramData:
        athlete = self._extract_athlete(workbook)
        file_info = parse_filename(filename) if filename else {}


        if file_info.get("athlete_name") and not athlete.name:
            athlete.name = file_info["athlete_name"]

        sb_name = self._find_strength_sheet(workbook) or "Strength Block"

        program = ProgramData(
            athlete=athlete,
            program_name=file_info.get("program_name"),
            program_number=file_info.get("program_number"),
            date_start=file_info.get("date_start"),
            date_end=file_info.get("date_end"),
            block_type=sb_name,
        )

        ws = workbook[sb_name]


        bands = _detect_week_bands(ws)


        for week_idx, band in enumerate(bands):
            week_number = week_idx + 1
            day_markers = _find_day_markers(ws, band["name_col"])

            if not day_markers:
                continue

            for i, marker in enumerate(day_markers):

                if i + 1 < len(day_markers):
                    end_row = day_markers[i + 1]["label_row"] - 1
                else:
                    end_row = min(marker["label_row"] + 30, ws.max_row)

                day_name, header_row = _find_day_name_and_header(
                    ws, marker["label_row"], band["name_col"], band
                )

                session_label = _find_session_label(ws, header_row, band["name_col"])

                pre_session = _extract_pre_session(
                    ws, marker["label_row"], band["name_col"] + band["sets"]
                )

                exercises = _extract_exercises(
                    ws, header_row, band["name_col"], band, end_row
                )

                session = SessionData(
                    info=SessionInfo(
                        week_number=week_number,
                        day_number=marker["day_number"],
                        day_name=day_name,
                        session_label=session_label,
                        nutrition_rating=pre_session["nutrition"],
                        stress_rating=pre_session["stress"],
                        sleep_rating=pre_session["sleep"],
                        fatigue_rating=pre_session["fatigue"],
                    ),
                    exercises=exercises,
                )
                program.sessions.append(session)


        program.weekly_volume = self._extract_weekly_volume(ws)


        program.daily_wellness = self._extract_daily_wellness(workbook)


        primary_days = _detect_primary_days(ws)
        program.primary_squat_day = primary_days["squat"]
        program.primary_bench_day = primary_days["bench"]
        program.primary_deadlift_day = primary_days["deadlift"]

        return program

    def _extract_athlete(self, workbook: openpyxl.Workbook) -> AthleteInfo:
        """Extract athlete metadata from the HomePage sheet."""
        ws = workbook["HomePage"]

        name = ""

        name_val = _cell_val(ws, 9, 2)
        if name_val:
            name = str(name_val).strip()


        sb_name = self._find_strength_sheet(workbook) or "Strength Block"
        if not name:
            sb = workbook[sb_name]
            for r in range(1, 10):
                v = _cell_val(sb, r, 1)
                if v and isinstance(v, str) and v.startswith("Client:"):
                    name = v.replace("Client:", "").strip()
                    break

        squat = _cell_val(ws, 4, 6)
        bench = _cell_val(ws, 4, 7)
        deadlift = _cell_val(ws, 4, 8)
        total = _cell_val(ws, 4, 9)
        goal = _cell_val(ws, 2, 12)
        meet_date = _cell_val(ws, 4, 12)

        return AthleteInfo(
            name=name,
            squat_max_lbs=float(squat) if isinstance(squat, (int, float)) else None,
            bench_max_lbs=float(bench) if isinstance(bench, (int, float)) else None,
            deadlift_max_lbs=float(deadlift) if isinstance(deadlift, (int, float)) else None,
            total_lbs=float(total) if isinstance(total, (int, float)) else None,
            goal=str(goal).strip() if goal else None,
            next_meet_date=str(meet_date).strip() if meet_date and meet_date != "TBD" else None,
        )

    def _extract_weekly_volume(self, ws: Worksheet) -> dict[int, dict[str, float | None]]:
        """Extract weekly volume summary rows from the bottom of the Strength Block."""
        volume: dict[int, dict[str, float | None]] = {}


        bands = _detect_week_bands(ws)
        for week_idx, band in enumerate(bands):
            week_num = week_idx + 1
            vol_col = band["name_col"] + band["volume"]
            label_col = vol_col - 1

            week_vol: dict[str, float | None] = {"squat": None, "bench": None, "deadlift": None}


            for r in range(75, min(120, ws.max_row + 1)):
                label = _cell_val(ws, r, label_col)
                val = _cell_val(ws, r, vol_col)

                if label and isinstance(label, str):
                    label_clean = label.strip().lower()
                    if isinstance(val, (int, float)):
                        if "squat" in label_clean:
                            week_vol["squat"] = float(val)
                        elif "bench" in label_clean:
                            week_vol["bench"] = float(val)
                        elif "deadlift" in label_clean or "dead" in label_clean:
                            week_vol["deadlift"] = float(val)

            if any(v is not None for v in week_vol.values()):
                volume[week_num] = week_vol

        return volume

    def _extract_daily_wellness(self, workbook: openpyxl.Workbook) -> list[WellnessEntry]:
        """Extract daily nutrition and bodyweight data from the HomePage tab.

        The HomePage has repeating weekly blocks with this layout:
            Row N,   col 11: "Nutritional Guide"
            Row N,   cols 13-19: Day headers (Monday-Sunday)
            Row N+1, col 12: "Goal", cols 13-19: "Actual"
            Row N+2, col 11: "Calories"  | col 12: goal | cols 13-19: daily actuals
            Row N+3, col 11: "Carbs"     | col 12: goal | cols 13-19: daily actuals
            Row N+4, col 11: "Fats"      | col 12: goal | cols 13-19: daily actuals
            Row N+5, col 11: "Protein"   | col 12: goal | cols 13-19: daily actuals
            Row N+6, col 11: "Water Intake"
            Row N+7, col 11: "AM Weight" | col 12: goal | cols 13-19: daily actuals
        """
        ws = workbook["HomePage"]
        entries: list[WellnessEntry] = []


        _DAY_COLS = {
            0: "Monday",
            1: "Tuesday",
            2: "Wednesday",
            3: "Thursday",
            4: "Friday",
            5: "Saturday",
            6: "Sunday",
        }


        guide_rows: list[int] = []
        for r in range(1, min(ws.max_row + 1, 100)):
            val = _cell_val(ws, r, 11)
            if val and isinstance(val, str) and val.strip() == "Nutritional Guide":
                guide_rows.append(r)

        for week_idx, guide_row in enumerate(guide_rows):
            week_number = week_idx + 1


            macro_data: dict[str, dict[str, float | None]] = {}
            weight_row_data: dict[int, float | None] = {}

            for r in range(guide_row + 2, guide_row + 10):
                label = _cell_val(ws, r, 11)
                if not label or not isinstance(label, str):
                    continue

                label_clean = label.strip().lower()

                if label_clean in ("calories", "carbs", "fats", "protein"):

                    goal_val = _cell_val(ws, r, 12)
                    goal_num = float(goal_val) if isinstance(goal_val, (int, float)) else None


                    for day_offset in range(7):
                        col = 13 + day_offset
                        actual_val = _cell_val(ws, r, col)
                        actual_num = float(actual_val) if isinstance(actual_val, (int, float)) else None

                        day_key = day_offset
                        if day_key not in macro_data:
                            macro_data[day_key] = {
                                "goal_calories": None, "goal_carbs": None,
                                "goal_fat": None, "goal_protein": None,
                                "actual_calories": None, "actual_carbs": None,
                                "actual_fat": None, "actual_protein": None,
                            }


                        if label_clean == "calories":
                            macro_data[day_key]["goal_calories"] = goal_num
                            macro_data[day_key]["actual_calories"] = actual_num
                        elif label_clean == "carbs":
                            macro_data[day_key]["goal_carbs"] = goal_num
                            macro_data[day_key]["actual_carbs"] = actual_num
                        elif label_clean == "fats":
                            macro_data[day_key]["goal_fat"] = goal_num
                            macro_data[day_key]["actual_fat"] = actual_num
                        elif label_clean == "protein":
                            macro_data[day_key]["goal_protein"] = goal_num
                            macro_data[day_key]["actual_protein"] = actual_num

                elif label_clean == "am weight":
                    for day_offset in range(7):
                        col = 13 + day_offset
                        wt_val = _cell_val(ws, r, col)
                        if isinstance(wt_val, (int, float)):
                            weight_row_data[day_offset] = float(wt_val)
                        else:
                            weight_row_data[day_offset] = None


            for day_offset in range(7):
                day_name = _DAY_COLS[day_offset]
                macros = macro_data.get(day_offset, {})
                bw = weight_row_data.get(day_offset)


                has_data = bw is not None or any(
                    macros.get(k) is not None
                    for k in ("actual_calories", "actual_carbs", "actual_fat",
                              "actual_protein", "goal_calories", "goal_carbs",
                              "goal_fat", "goal_protein")
                )

                if has_data:
                    entries.append(WellnessEntry(
                        week_number=week_number,
                        day_of_week=day_name,
                        goal_calories=macros.get("goal_calories"),
                        goal_protein=macros.get("goal_protein"),
                        goal_carbs=macros.get("goal_carbs"),
                        goal_fat=macros.get("goal_fat"),
                        actual_calories=macros.get("actual_calories"),
                        actual_protein=macros.get("actual_protein"),
                        actual_carbs=macros.get("actual_carbs"),
                        actual_fat=macros.get("actual_fat"),
                        bodyweight_lbs=bw,
                    ))

        return entries
